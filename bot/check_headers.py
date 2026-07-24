import openpyxl

LOCAL_FILE = r"C:\Users\Yisus\Desktop\Atajos Globales\Plan de cosecha 2026 Test.xlsx"
wb = openpyxl.load_workbook(LOCAL_FILE, data_only=True)

sheet_name = [s for s in wb.sheetnames if s.startswith("P Cosecha ")][-1]
ws = wb[sheet_name]

print(f"--- HEADERS OF {sheet_name} ---")
for col in range(1, 40):
    h3 = ws.cell(row=3, column=col).value
    h4 = ws.cell(row=4, column=col).value
    h5 = ws.cell(row=5, column=col).value
    print(f"Col {col}: row3={h3}, row4={h4}, row5={h5}")

wb.close()
