from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from excel_range_sync import CHUNK_SIZE, clean_value, graph_request
from sharepoint_sync import GRAPH_URL, graph_headers


SHEET_NAME = "Inventario"
SOURCE_HEADERS = (
    "flor",
    "color",
    "tallos_por_ramo",
    "tallos_en_existencia",
)


def read_inventory_rows(report_path: Path) -> list[list[str | int | float | bool]]:
    workbook = load_workbook(report_path, read_only=True, data_only=False)
    try:
        sheet = workbook.active
        headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
        missing = [header for header in SOURCE_HEADERS if header not in headers]
        if missing:
            raise RuntimeError(
                f"Faltan columnas en el reporte de Inventario: {', '.join(missing)}."
            )
        indexes = [headers.index(header) + 1 for header in SOURCE_HEADERS]
        rows = [
            [clean_value(sheet.cell(row=row, column=column).value) for column in indexes]
            for row in range(2, sheet.max_row + 1)
        ]
    finally:
        workbook.close()
    return rows


def worksheet_range_url(
    workbook_url: str, start_column: str, start_row: int, end_column: str, end_row: int
) -> str:
    address = f"{start_column}{start_row}:{end_column}{end_row}"
    return (
        f"{workbook_url}/worksheets/{SHEET_NAME}/"
        f"range(address='{address}')"
    )


def patch_values(
    workbook_url: str,
    headers: dict[str, str],
    start_column: str,
    end_column: str,
    start_row: int,
    values: list[list[Any]],
) -> None:
    if not values:
        return
    graph_request(
        "PATCH",
        worksheet_range_url(
            workbook_url,
            start_column,
            start_row,
            end_column,
            start_row + len(values) - 1,
        ),
        headers,
        json={"values": values},
    )


def patch_formulas(
    workbook_url: str,
    headers: dict[str, str],
    column: str,
    start_row: int,
    formulas: list[list[str]],
) -> None:
    if not formulas:
        return
    graph_request(
        "PATCH",
        worksheet_range_url(
            workbook_url,
            column,
            start_row,
            column,
            start_row + len(formulas) - 1,
        ),
        headers,
        json={"formulas": formulas},
    )


def update_inventory_range(token: str, item: dict, report_path: Path) -> int:
    rows = read_inventory_rows(report_path)
    drive_id = item["parentReference"]["driveId"]
    workbook_url = f"{GRAPH_URL}/drives/{drive_id}/items/{item['id']}/workbook"
    headers = {**graph_headers(token), "Content-Type": "application/json"}
    session = graph_request(
        "POST",
        f"{workbook_url}/createSession",
        headers,
        json={"persistChanges": True},
        timeout=60,
    ).json()
    session_headers = {**headers, "workbook-session-id": session["id"]}

    try:
        first_headers = graph_request(
            "GET",
            worksheet_range_url(workbook_url, "A", 1, "B", 1),
            session_headers,
            timeout=60,
        ).json().get("values", [[]])[0]
        second_headers = graph_request(
            "GET",
            worksheet_range_url(workbook_url, "D", 1, "E", 1),
            session_headers,
            timeout=60,
        ).json().get("values", [[]])[0]
        if [str(value).strip().lower() for value in first_headers] != ["flor", "color"]:
            raise RuntimeError("Los encabezados A:B de Inventario no coinciden.")
        if [str(value).strip().lower() for value in second_headers] != [
            "tallos_por_ramo",
            "tallos_en_existencia",
        ]:
            raise RuntimeError("Los encabezados D:E de Inventario no coinciden.")

        used_range = graph_request(
            "GET",
            f"{workbook_url}/worksheets/{SHEET_NAME}/usedRange",
            session_headers,
            timeout=60,
        ).json()
        current_last_row = max(2, int(used_range.get("rowCount") or 1))
        new_last_row = len(rows) + 1
        clear_last_row = max(current_last_row, new_last_row)

        for start_row in range(2, clear_last_row + 1, CHUNK_SIZE):
            count = min(CHUNK_SIZE, clear_last_row - start_row + 1)
            patch_values(
                workbook_url,
                session_headers,
                "A",
                "B",
                start_row,
                [["", ""] for _ in range(count)],
            )
            patch_values(
                workbook_url,
                session_headers,
                "D",
                "E",
                start_row,
                [["", ""] for _ in range(count)],
            )

        for offset in range(0, len(rows), CHUNK_SIZE):
            batch = rows[offset : offset + CHUNK_SIZE]
            start_row = offset + 2
            patch_values(
                workbook_url,
                session_headers,
                "A",
                "B",
                start_row,
                [[row[0], row[1]] for row in batch],
            )
            patch_values(
                workbook_url,
                session_headers,
                "D",
                "E",
                start_row,
                [[row[2], row[3]] for row in batch],
            )
            patch_formulas(
                workbook_url,
                session_headers,
                "C",
                start_row,
                [[f'=+A{row_number} & " " & B{row_number}'] for row_number in range(start_row, start_row + len(batch))],
            )
            patch_formulas(
                workbook_url,
                session_headers,
                "F",
                start_row,
                [[f'=+A{row_number} & " " & B{row_number} & "X" & " " & D{row_number}'] for row_number in range(start_row, start_row + len(batch))],
            )

        if rows:
            verification = graph_request(
                "GET",
                worksheet_range_url(workbook_url, "A", new_last_row, "E", new_last_row),
                session_headers,
                timeout=60,
            ).json().get("values", [])
            if not verification:
                raise RuntimeError("Excel Online no devolvió la última fila de Inventario.")

        print(
            f"INVENTARIO_RANGE_UPDATE_OK hoja={SHEET_NAME} "
            f"columnas=A,B,D,E filas_datos={len(rows)} libro_no_reemplazado=true",
            flush=True,
        )
        return len(rows)
    finally:
        try:
            graph_request(
                "POST",
                f"{workbook_url}/closeSession",
                session_headers,
                timeout=30,
            )
        except Exception as exc:
            print(f"Aviso al cerrar la sesión de Inventario: {exc}", flush=True)
