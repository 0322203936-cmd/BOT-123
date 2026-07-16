import math
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

from sharepoint_sync import GRAPH_URL, graph_headers


SHEET_NAME = "PegarData"
TABLE_NAME = "Table2"
DATA_COLUMNS = 18
TABLE_COLUMNS = 23
CHUNK_SIZE = 200
RETRYABLE_STATUS = {409, 429, 502, 503, 504}
CALCULATED_HEADERS = [
    "Flor Tallos",
    "TxR2",
    "Flor Color2",
    "Flor Color 3",
    "Flor Color 4",
]


def clean_value(value: Any) -> str | int | float | bool:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and not math.isfinite(value):
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def read_report_rows(report_path: Path) -> list[list[str | int | float | bool]]:
    workbook = load_workbook(report_path, read_only=True, data_only=False)
    try:
        sheet = workbook.active
        if sheet.max_column != DATA_COLUMNS:
            raise RuntimeError(
                f"El reporte tiene {sheet.max_column} columnas; se esperaban 18 (A:R)."
            )
        rows = [
            [clean_value(cell.value) for cell in row]
            for row in sheet.iter_rows(min_col=1, max_col=DATA_COLUMNS)
        ]
    finally:
        workbook.close()

    if not rows:
        raise RuntimeError("El reporte no contiene filas.")
    if any(len(row) != DATA_COLUMNS for row in rows):
        raise RuntimeError("El reporte contiene filas con una cantidad inesperada de columnas.")
    return rows


def graph_request(
    method: str,
    url: str,
    headers: dict[str, str],
    *,
    json: dict[str, Any] | None = None,
    timeout: int = 120,
) -> requests.Response:
    for attempt in range(1, 7):
        response = requests.request(
            method, url, headers=headers, json=json, timeout=timeout
        )
        if response.status_code in RETRYABLE_STATUS and attempt < 6:
            delay = int(response.headers.get("Retry-After", "0") or 0) or min(
                5 * attempt, 30
            )
            print(
                f"Excel Online respondio HTTP {response.status_code}; "
                f"reintento {attempt}/5 en {delay} segundos...",
                flush=True,
            )
            time.sleep(delay)
            continue
        if not response.ok:
            try:
                detail = response.json().get("error", {}).get("message", "")
            except ValueError:
                detail = response.text
            raise RuntimeError(
                f"Microsoft Graph respondio {response.status_code}: {detail}"
            )
        return response
    raise RuntimeError("Microsoft Graph no respondio despues de los reintentos.")


def range_url(workbook_url: str, start_row: int, end_row: int) -> str:
    address = f"A{start_row}:R{end_row}"
    return (
        f"{workbook_url}/worksheets/{SHEET_NAME}/"
        f"range(address='{address}')"
    )


def calculated_range_url(workbook_url: str, start_row: int, end_row: int) -> str:
    address = f"S{start_row}:W{end_row}"
    return (
        f"{workbook_url}/worksheets/{SHEET_NAME}/"
        f"range(address='{address}')"
    )


def calculated_formulas(row: int) -> list[str]:
    return [
        '=+Table2[[#This Row],[FLOR]]&" "&"X"&" " & Table2[[#This Row],[txr_orden]]',
        "=Table2[[#This Row],[txr_orden]]",
        '=+Table2[[#This Row],[FLOR]] & " " &Table2[[#This Row],[Flor Color]]',
        '=+Table2[[#This Row],[FLOR]] & " " & Table2[[#This Row],[Flor Color]]&"X"&" " & Table2[[#This Row],[TxR2]]',
        f'=E{row} & " " & " X " & T{row}',
    ]


def patch_calculated_rows(
    workbook_url: str,
    headers: dict[str, str],
    start_row: int,
    count: int,
) -> None:
    if count < 1:
        return
    end_row = start_row + count - 1
    formulas = [calculated_formulas(row) for row in range(start_row, end_row + 1)]
    graph_request(
        "PATCH",
        calculated_range_url(workbook_url, start_row, end_row),
        headers,
        json={"formulas": formulas},
        timeout=120,
    )


def verify_calculated_rows(
    workbook_url: str,
    headers: dict[str, str],
    start_row: int,
    count: int,
) -> None:
    if count < 1:
        return
    end_row = start_row + count - 1
    formulas = graph_request(
        "GET",
        calculated_range_url(workbook_url, start_row, end_row),
        headers,
        timeout=120,
    ).json().get("formulas", [])
    if len(formulas) != count or any(
        len(formula_row) != len(CALCULATED_HEADERS)
        or any(
            not isinstance(formula, str) or not formula.startswith("=")
            for formula in formula_row
        )
        for formula_row in formulas
    ):
        raise RuntimeError(
            f"Las formulas S:W no se propagaron completamente en las filas "
            f"{start_row}:{end_row}."
        )


def patch_rows(
    workbook_url: str,
    headers: dict[str, str],
    start_row: int,
    rows: list[list[str | int | float | bool]],
) -> None:
    if not rows:
        return
    end_row = start_row + len(rows) - 1
    graph_request(
        "PATCH",
        range_url(workbook_url, start_row, end_row),
        headers,
        json={"values": rows},
    )


def add_table_rows(
    workbook_url: str, headers: dict[str, str], count: int
) -> None:
    remaining = count
    while remaining > 0:
        batch = min(CHUNK_SIZE, remaining)
        graph_request(
            "POST",
            f"{workbook_url}/tables/{TABLE_NAME}/rows/add",
            headers,
            json={"index": None, "values": [[""] * TABLE_COLUMNS for _ in range(batch)]},
        )
        remaining -= batch


def update_pegar_data_range(token: str, item: dict, report_path: Path) -> int:
    rows = read_report_rows(report_path)
    report_headers = rows[0]
    report_data = rows[1:]

    drive_id = item["parentReference"]["driveId"]
    workbook_url = f"{GRAPH_URL}/drives/{drive_id}/items/{item['id']}/workbook"
    headers = {**graph_headers(token), "Content-Type": "application/json"}
    session = graph_request(
        "POST",
        f"{workbook_url}/createSession",
        headers,
        json={"persistChanges": True},
        timeout=60,
    ).json()
    session_headers = {**headers, "workbook-session-id": session["id"]}

    try:
        current_headers = graph_request(
            "GET",
            range_url(workbook_url, 1, 1),
            session_headers,
            timeout=60,
        ).json().get("values", [[]])[0]
        if current_headers != report_headers:
            raise RuntimeError(
                "Los encabezados A:R de PegarData no coinciden con el reporte; "
                "se cancelo la escritura para proteger el libro."
            )

        calculated_headers = graph_request(
            "GET",
            calculated_range_url(workbook_url, 1, 1),
            session_headers,
            timeout=60,
        ).json().get("values", [[]])[0]
        if calculated_headers != CALCULATED_HEADERS:
            raise RuntimeError(
                "Los encabezados S:W de PegarData no coinciden; "
                "se cancelo la escritura para proteger las formulas."
            )

        table = graph_request(
            "GET",
            f"{workbook_url}/tables/{TABLE_NAME}/range",
            session_headers,
            timeout=60,
        ).json()
        current_table_rows = int(table.get("rowCount") or 0)
        if current_table_rows < 1:
            raise RuntimeError("Table2 no devolvio una cantidad valida de filas.")

        required_table_rows = len(rows)
        if required_table_rows > current_table_rows:
            add_table_rows(
                workbook_url,
                session_headers,
                required_table_rows - current_table_rows,
            )
            current_table_rows = required_table_rows

        # Conserva los encabezados y limpia solamente los datos A:R existentes.
        blank_row = [""] * DATA_COLUMNS
        for start_row in range(2, current_table_rows + 1, CHUNK_SIZE):
            count = min(CHUNK_SIZE, current_table_rows - start_row + 1)
            patch_rows(
                workbook_url,
                session_headers,
                start_row,
                [blank_row[:] for _ in range(count)],
            )

        for offset in range(0, len(report_data), CHUNK_SIZE):
            patch_rows(
                workbook_url,
                session_headers,
                offset + 2,
                report_data[offset : offset + CHUNK_SIZE],
            )

        for start_row in range(2, len(rows) + 1, CHUNK_SIZE):
            count = min(CHUNK_SIZE, len(rows) - start_row + 1)
            patch_calculated_rows(
                workbook_url,
                session_headers,
                start_row,
                count,
            )
            verify_calculated_rows(
                workbook_url,
                session_headers,
                start_row,
                count,
            )

        last_row = max(len(rows), 1)
        final_row = graph_request(
            "GET",
            range_url(workbook_url, last_row, last_row),
            session_headers,
            timeout=60,
        ).json().get("values", [])
        if report_data and not final_row:
            raise RuntimeError("Excel Online no devolvio la ultima fila escrita.")

        print(
            f"EXCEL_RANGE_UPDATE_OK hoja={SHEET_NAME} rango=A1:R{last_row} "
            f"formulas=S2:W{last_row} filas_datos={len(report_data)} "
            "libro_no_reemplazado=true",
            flush=True,
        )
        return len(report_data)
    finally:
        try:
            graph_request(
                "POST",
                f"{workbook_url}/closeSession",
                session_headers,
                timeout=30,
            )
        except Exception as exc:
            print(f"Aviso al cerrar la sesion de Excel: {exc}", flush=True)
