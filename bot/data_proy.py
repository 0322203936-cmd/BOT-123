import os
import re
import sys
from pathlib import Path
import openpyxl

from sharepoint_sync import (
    GRAPH_URL,
    graph_headers,
    graph_token,
    resolve_sharepoint_item_by_url,
    download_sharepoint_file,
)
from excel_range_sync import graph_request

REQ_PROY_URL = "https://pacificafarms.sharepoint.com/:x:/r/sites/requerimientovsproyeccion/_layouts/15/Doc.aspx?sourcedoc=%7B277A76AA-508A-47F8-8A4A-F19D46660D65%7D&file=Requerimiento%20vs%20proyeccion%20Test.xlsm&action=default&mobileredirect=true"
PLAN_COSECHA_URL = "https://pacificafarms.sharepoint.com/:x:/r/sites/requerimientovsproyeccion/_layouts/15/Doc.aspx?sourcedoc=%7B0A3464AB-7BD8-400A-A0E6-5BC92E23CE3E%7D&file=Plan%20de%20cosecha%202026%20Test.xlsx&action=default&mobileredirect=true"

def col_letter_to_index(letter):
    idx = 0
    for char in letter.upper():
        idx = idx * 26 + (ord(char) - ord('A') + 1)
    return idx - 1

def main():
    print("Obteniendo token de Microsoft Graph...")
    token = graph_token()
    
    print("Resolviendo archivos en SharePoint...")
    item_req = resolve_sharepoint_item_by_url(token, REQ_PROY_URL)
    item_plan = resolve_sharepoint_item_by_url(token, PLAN_COSECHA_URL)
    
    print("Abriendo sesión en vivo de Plan de Cosecha...")
    plan_drive_id = item_plan["parentReference"]["driveId"]
    plan_workbook_url = f"{GRAPH_URL}/drives/{plan_drive_id}/items/{item_plan['id']}/workbook"
    headers = {**graph_headers(token), "Content-Type": "application/json"}
    
    plan_session_res = graph_request(
        "POST",
        f"{plan_workbook_url}/createSession",
        headers,
        json={"persistChanges": False},
        timeout=60,
    ).json()
    plan_session_headers = {**headers, "workbook-session-id": plan_session_res["id"]}
    
    try:
        worksheets_data = graph_request("GET", f"{plan_workbook_url}/worksheets", plan_session_headers).json()
        sheetnames = [ws["name"] for ws in worksheets_data.get("value", [])]
        
        cosecha_sheets = [s for s in sheetnames if str(s).startswith("P Cosecha ")]
        if not cosecha_sheets:
            print("Error: No se encontraron hojas de 'P Cosecha'")
            sys.exit(1)
            
        latest_sheet_name = sorted(cosecha_sheets)[-1]
        print(f"Hoja detectada (en vivo): {latest_sheet_name}")
        
        from urllib.parse import quote
        quoted_sheet = quote(latest_sheet_name)
        
        print("Obteniendo datos en vivo...")
        range_data = graph_request("GET", f"{plan_workbook_url}/worksheets('{quoted_sheet}')/range(address='A1:X1000')", plan_session_headers).json()
        plan_values = range_data.get("values", [])
        
        import re
        def get_week(col):
            if len(plan_values) <= 3: return ""
            if col - 1 >= len(plan_values[3]): return ""
            val = plan_values[3][col - 1]
            if val is None or str(val).strip() == "": return ""
            nums = re.findall(r'\d+', str(val))
            return int(nums[0]) if nums else str(val).strip()
            
        weeks = [
            get_week(6),
            get_week(18),
            get_week(19),
            get_week(20),
            get_week(21),
            get_week(22),
            get_week(23),
            get_week(24)
        ]
        print(f"Semanas detectadas desde los encabezados: {weeks}")
        print(f"Semanas a procesar: {weeks}")
        
        flowers_data = []
        for row_idx in range(4, len(plan_values)):
            row_data = plan_values[row_idx]
            if not row_data: continue
            
            flor_val = row_data[0]
            if flor_val is not None and str(flor_val).strip().lower() == "total":
                break
            if not flor_val:
                continue
                
            flor_real = row_data[3] if len(row_data) > 3 else None
            color_real = row_data[4] if len(row_data) > 4 else None
            
            if not flor_real:
                continue
                
            qty_30 = row_data[5] if len(row_data) > 5 and row_data[5] is not None else 0
            qty_31 = row_data[17] if len(row_data) > 17 and row_data[17] is not None else 0
            qty_32 = row_data[18] if len(row_data) > 18 and row_data[18] is not None else 0
            qty_33 = row_data[19] if len(row_data) > 19 and row_data[19] is not None else 0
            qty_34 = row_data[20] if len(row_data) > 20 and row_data[20] is not None else 0
            qty_35 = row_data[21] if len(row_data) > 21 and row_data[21] is not None else 0
            qty_36 = row_data[22] if len(row_data) > 22 and row_data[22] is not None else 0
            qty_37 = row_data[23] if len(row_data) > 23 and row_data[23] is not None else 0
            
            flowers_data.append({
                "flor": flor_real,
                "color": color_real,
                "qtys": [qty_30, qty_31, qty_32, qty_33, qty_34, qty_35, qty_36, qty_37]
            })
            
        print(f"Se extrajeron {len(flowers_data)} flores.")
    finally:
        try:
            graph_request("POST", f"{plan_workbook_url}/closeSession", plan_session_headers, timeout=30)
        except:
            pass

    if os.environ.get("SHAREPOINT_UPLOAD", "true").lower() not in {"1", "true", "yes", "si", "sí"}:
        print("Modo prueba: SHAREPOINT_UPLOAD está apagado. Deteniendo ejecución.")
        return

    print("\nIniciando Edición en Vivo (Live Patching) en Requerimiento vs proyeccion...")
    drive_id = item_req["parentReference"]["driveId"]
    workbook_url = f"{GRAPH_URL}/drives/{drive_id}/items/{item_req['id']}/workbook"
    headers = {**graph_headers(token), "Content-Type": "application/json"}
    
    print("Abriendo sesión en Excel Online...")
    session = graph_request(
        "POST",
        f"{workbook_url}/createSession",
        headers,
        json={"persistChanges": True},
        timeout=60,
    ).json()
    session_headers = {**headers, "workbook-session-id": session["id"]}
    
    try:
        print("Obteniendo área de trabajo (usedRange)...")
        used_range_res = graph_request(
            "GET", 
            f"{workbook_url}/worksheets/DataProy/usedRange", 
            session_headers,
            timeout=120
        ).json()
        
        address = used_range_res.get("address", "")
        values = used_range_res.get("values", [])
        
        match = re.search(r'!([A-Za-z]+)(\d+)', address)
        start_col_str = match.group(1).upper() if match else "A"
        start_row = int(match.group(2)) if match else 1
        
        start_col_idx = col_letter_to_index(start_col_str)
        col_h_rel = col_letter_to_index("H") - start_col_idx
        col_s_rel = col_letter_to_index("S") - start_col_idx
        
        rows_by_week = {w: [] for w in weeks}
        rows_to_clear = []
        
        corte_start_excel_row = None
        seen_compra = False
        
        for idx, row_data in enumerate(values):
            if len(row_data) > col_h_rel and col_h_rel >= 0:
                desc = str(row_data[col_h_rel]).strip().upper()
            else:
                desc = ""
                
            if desc == "COMPRA":
                seen_compra = True
            elif seen_compra and desc != "COMPRA":
                corte_start_excel_row = start_row + idx
                break
                
        if not corte_start_excel_row:
            if seen_compra:
                corte_start_excel_row = start_row + len(values)
            else:
                # Si no se encuentra COMPRA, lo ponemos al final del rango usado
                corte_start_excel_row = start_row + len(values)
                
        num_flowers = len(flowers_data)
        total_corte_rows_needed = num_flowers * len(weeks)
        
        # Asignar filas exactas
        for i in range(total_corte_rows_needed):
            week_idx = i // num_flowers
            if week_idx < len(weeks):
                current_week = weeks[week_idx]
                rows_by_week[current_week].append(corte_start_excel_row + i)
                
        # Buscar filas residuales viejas que limpiar
        end_of_corte_excel_row = corte_start_excel_row + total_corte_rows_needed
        for idx, row_data in enumerate(values):
            excel_row = start_row + idx
            if excel_row >= end_of_corte_excel_row:
                if len(row_data) > col_h_rel and col_h_rel >= 0:
                    desc = str(row_data[col_h_rel]).strip().upper()
                    if desc == "CORTE":
                        rows_to_clear.append(excel_row)
                    
        for i, week in enumerate(weeks):
            target_rows = sorted(rows_by_week[week])
            if not target_rows:
                print(f"Advertencia: No se encontraron filas CORTE para semana {week}.")
                continue
                
            blocks = []
            current_block = []
            for r in target_rows:
                if not current_block or r == current_block[-1] + 1:
                    current_block.append(r)
                else:
                    blocks.append(current_block)
                    current_block = [r]
            if current_block:
                blocks.append(current_block)
                
            flower_idx = 0
            for block in blocks:
                start_r = block[0]
                end_r = block[-1]
                count = len(block)
                
                flor_color_desc_values = []
                tallos_values = []
                semana_values = []
                
                for _ in range(count):
                    if flower_idx < len(flowers_data):
                        item = flowers_data[flower_idx]
                        flor_color_desc_values.append([str(item["flor"]) if item["flor"] else "", str(item["color"]) if item["color"] else "", "CORTE"])
                        tallos_values.append([str(item["qtys"][i]) if item["qtys"][i] is not None else "0"])
                    else:
                        flor_color_desc_values.append(["", "", "CORTE"])
                        tallos_values.append([""])
                    semana_values.append([str(week)])
                    flower_idx += 1
                    
                print(f"Semana {week}: Escribiendo bloque filas {start_r}:{end_r}...")
                address_fh = f"F{start_r}:H{end_r}"
                empty_fh = [["", "", ""] for _ in range(count)]
                graph_request("PATCH", f"{workbook_url}/worksheets/DataProy/range(address='{address_fh}')", session_headers, json={"values": empty_fh})
                graph_request(
                    "PATCH", 
                    f"{workbook_url}/worksheets/DataProy/range(address='{address_fh}')", 
                    session_headers, 
                    json={"values": flor_color_desc_values}
                )
                
                address_o = f"O{start_r}:O{end_r}"
                empty_o = [[""] for _ in range(count)]
                graph_request("PATCH", f"{workbook_url}/worksheets/DataProy/range(address='{address_o}')", session_headers, json={"values": empty_o})
                graph_request(
                    "PATCH", 
                    f"{workbook_url}/worksheets/DataProy/range(address='{address_o}')", 
                    session_headers, 
                    json={"values": tallos_values}
                )
                
                address_s = f"S{start_r}:S{end_r}"
                empty_s = [[""] for _ in range(count)]
                graph_request("PATCH", f"{workbook_url}/worksheets/DataProy/range(address='{address_s}')", session_headers, json={"values": empty_s})
                graph_request(
                    "PATCH", 
                    f"{workbook_url}/worksheets/DataProy/range(address='{address_s}')", 
                    session_headers, 
                    json={"values": semana_values}
                )

        if rows_to_clear:
            print(f"Limpiando {len(rows_to_clear)} filas residuales de CORTE de ejecuciones anteriores...")
            clear_blocks = []
            current_clear_block = []
            for r in sorted(rows_to_clear):
                if not current_clear_block:
                    current_clear_block.append(r)
                elif r == current_clear_block[-1] + 1:
                    current_clear_block.append(r)
                else:
                    clear_blocks.append(current_clear_block)
                    current_clear_block = [r]
            if current_clear_block:
                clear_blocks.append(current_clear_block)
                
            for block in clear_blocks:
                start_r = block[0]
                end_r = block[-1]
                
                print(f"Limpiando bloque {start_r}:{end_r}")
                
                address_fh = f"F{start_r}:H{end_r}"
                graph_request("POST", f"{workbook_url}/worksheets/DataProy/range(address='{address_fh}')/clear", session_headers, json={"applyTo": "contents"})
                
                address_o = f"O{start_r}:O{end_r}"
                graph_request("POST", f"{workbook_url}/worksheets/DataProy/range(address='{address_o}')/clear", session_headers, json={"applyTo": "contents"})
                
                address_s = f"S{start_r}:S{end_r}"
                graph_request("POST", f"{workbook_url}/worksheets/DataProy/range(address='{address_s}')/clear", session_headers, json={"applyTo": "contents"})

        print("Edición en vivo finalizada con éxito.")
    finally:
        print("Cerrando sesión de Excel Online...")
        try:
            graph_request(
                "POST",
                f"{workbook_url}/closeSession",
                session_headers,
                timeout=30,
            )
        except Exception as exc:
            print(f"Aviso al cerrar la sesion: {exc}")

if __name__ == "__main__":
    main()
