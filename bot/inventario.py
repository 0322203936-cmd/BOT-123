import os
import re
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright


POSCO_URL = "http://3.132.9.174/Posco/"
ARTIFACTS_DIR = Path("artifacts/inventario")
CAPTURES_DIR = ARTIFACTS_DIR / "capturas"
REPORTS_DIR = ARTIFACTS_DIR / "reportes"


def required_secret(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Falta configurar el secreto {name}.")
    return value


def env_flag(name: str, default: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "si", "sí"}


def capture(page: Page, filename: str) -> None:
    CAPTURES_DIR.mkdir(parents=True, exist_ok=True)
    page.screenshot(path=str(CAPTURES_DIR / filename), full_page=True)


def click_exact_text(page: Page, text: str) -> None:
    pattern = re.compile(rf"^\s*{re.escape(text)}\s*$", re.IGNORECASE)
    candidates = [
        page.get_by_role("button", name=pattern),
        page.get_by_role("link", name=pattern),
        page.get_by_text(pattern),
    ]
    for candidate in candidates:
        try:
            candidate.first.wait_for(state="visible", timeout=3_000)
            candidate.first.click(timeout=5_000)
            return
        except PlaywrightTimeoutError:
            continue
    raise RuntimeError(f"No se encontró el elemento visible: {text}.")


def open_inventory_general(page: Page) -> None:
    click_exact_text(page, "Inventario")
    page.wait_for_timeout(700)
    capture(page, "03_menu_inventario.png")
    click_exact_text(page, "Inventario General")
    try:
        page.wait_for_url("**/#/list-inventario-gral", timeout=30_000)
    except PlaywrightTimeoutError:
        print(f"Aviso: URL actual después de navegar: {page.url}", flush=True)
    page.wait_for_timeout(15_000)


def export_inventory(page: Page) -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    with page.expect_download(timeout=60_000) as download_info:
        click_exact_text(page, "Exportar")
    download = download_info.value
    filename = Path(download.suggested_filename).name or "inventario_general.xlsx"
    destination = REPORTS_DIR / filename
    download.save_as(str(destination))
    return destination


def inspect_report(report_path: Path) -> None:
    workbook = load_workbook(report_path, read_only=True, data_only=False)
    try:
        sheet = workbook.active
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        print(
            f"INVENTARIO_REPORT_OK archivo={report_path} filas={sheet.max_row} "
            f"columnas={sheet.max_column} encabezados={headers}",
            flush=True,
        )
    finally:
        workbook.close()


def run() -> None:
    user = required_secret("POSCO_USER")
    password = required_secret("POSCO_PASSWORD")

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context(viewport={"width": 1920, "height": 1080})
        page = context.new_page()
        try:
            print("Abriendo Posco...", flush=True)
            page.goto(POSCO_URL, wait_until="networkidle", timeout=60_000)
            capture(page, "01_login.png")

            page.locator(
                'input[placeholder*="usuario@email.com" i], input[type="text"]'
            ).first.fill(user)
            page.locator(
                'input[placeholder*="Password" i], input[type="password"]'
            ).first.fill(password)
            page.get_by_role("button", name="Iniciar Sesión").click(timeout=15_000)
            page.wait_for_load_state("networkidle", timeout=60_000)
            page.wait_for_timeout(2_000)
            capture(page, "02_dashboard.png")

            print("Abriendo Inventario > Inventario General...", flush=True)
            open_inventory_general(page)
            capture(page, "04_inventario_general.png")

            print("Exportando Inventario General...", flush=True)
            report = export_inventory(page)
            inspect_report(report)
            capture(page, "05_exportacion_completada.png")

            if env_flag("INVENTARIO_UPLOAD", default=False):
                raise RuntimeError(
                    "La escritura de Inventario seguirá desactivada hasta validar los encabezados."
                )
            print("Modo exploratorio: SharePoint no fue modificado.", flush=True)
        except Exception:
            capture(page, "99_error.png")
            raise
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    run()
