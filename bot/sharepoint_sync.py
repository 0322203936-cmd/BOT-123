import base64
import hashlib
import os
import time
from pathlib import Path
from zipfile import ZipFile

import requests
from openpyxl import load_workbook
from xlsm_patcher import patch_xlsm


GRAPH_URL = "https://graph.microsoft.com/v1.0"
SHAREPOINT_FILE_URL = (
    "https://pacificafarms.sharepoint.com/:x:/r/sites/"
    "requerimientovsproyeccion/_layouts/15/Doc.aspx?"
    "sourcedoc=%7BB4151F11-DDFD-48DC-8283-F6D684396F87%7D&"
    "file=Reunion%201-2-3%20Copia%20IRENE.xlsm&action=default&mobileredirect=true"
)
SHAREPOINT_DIR = Path("artifacts/sharepoint")
LOCK_RETRY_SECONDS = 20
LOCK_RETRY_ATTEMPTS = 46
VERSION_RETRY_ATTEMPTS = 10


class SharePointVersionConflict(RuntimeError):
    """El libro cambió entre la descarga y la subida."""


def required_secret(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Falta configurar el secreto {name}.")
    return value


def graph_token() -> str:
    tenant_id = required_secret("SHAREPOINT_TENANT_ID")
    response = requests.post(
        f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token",
        data={
            "client_id": required_secret("SHAREPOINT_CLIENT_ID"),
            "client_secret": required_secret("SHAREPOINT_CLIENT_SECRET"),
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        },
        timeout=30,
    )
    response.raise_for_status()
    return response.json()["access_token"]


def graph_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def share_id(url: str) -> str:
    encoded = base64.urlsafe_b64encode(url.encode("utf-8")).decode("ascii")
    return f"u!{encoded.rstrip('=')}"


def resolve_sharepoint_item(token: str) -> dict:
    response = requests.get(
        f"{GRAPH_URL}/shares/{share_id(SHAREPOINT_FILE_URL)}/driveItem",
        headers=graph_headers(token),
        timeout=30,
    )
    response.raise_for_status()
    item = response.json()
    if item.get("name", "").lower() != "reunion 1-2-3 copia irene.xlsm":
        raise RuntimeError(f"SharePoint devolvió un archivo inesperado: {item.get('name')}.")
    return item


def resolve_sharepoint_item_by_url(token: str, url: str) -> dict:
    response = requests.get(
        f"{GRAPH_URL}/shares/{share_id(url)}/driveItem",
        headers=graph_headers(token),
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def download_sharepoint_workbook(token: str, item: dict) -> Path:
    drive_id = item["parentReference"]["driveId"]
    item_id = item["id"]
    response = requests.get(
        f"{GRAPH_URL}/drives/{drive_id}/items/{item_id}/content",
        headers=graph_headers(token),
        timeout=120,
    )
    response.raise_for_status()
    SHAREPOINT_DIR.mkdir(parents=True, exist_ok=True)
    destination = SHAREPOINT_DIR / "Reunion 1-2-3 Test_original.xlsm"
    destination.write_bytes(response.content)
    print(f"Libro de SharePoint descargado: {destination}")
    return destination


def download_sharepoint_file(token: str, item: dict, dest_filename: str) -> Path:
    drive_id = item["parentReference"]["driveId"]
    item_id = item["id"]
    response = requests.get(
        f"{GRAPH_URL}/drives/{drive_id}/items/{item_id}/content",
        headers=graph_headers(token),
        timeout=120,
    )
    response.raise_for_status()
    SHAREPOINT_DIR.mkdir(parents=True, exist_ok=True)
    destination = SHAREPOINT_DIR / dest_filename
    destination.write_bytes(response.content)
    print(f"Archivo de SharePoint descargado: {destination}")
    return destination


def vba_hash(path: Path) -> str:
    with ZipFile(path) as archive:
        try:
            content = archive.read("xl/vbaProject.bin")
        except KeyError as exc:
            raise RuntimeError("El libro XLSM no contiene xl/vbaProject.bin.") from exc
    return hashlib.sha256(content).hexdigest()


def replace_pegar_data(report_path: Path, workbook_path: Path) -> Path:
    original_vba = vba_hash(workbook_path)
    destination = SHAREPOINT_DIR / "Reunion 1-2-3 Test_actualizado.xlsm"
    patch_xlsm(report_path, workbook_path, destination)

    if vba_hash(destination) != original_vba:
        raise RuntimeError("La validación detectó un cambio en el proyecto VBA.")

    source_book = load_workbook(report_path, read_only=True, data_only=False)
    source = source_book.active
    verification_book = load_workbook(
        destination, keep_vba=True, keep_links=True, data_only=False
    )
    verification = verification_book["PegarData"]
    expected_table_ref = f"A1:W{source.max_row}"
    if verification.tables["Table2"].ref != expected_table_ref:
        raise RuntimeError("Table2 no fue ampliada al nuevo rango de datos.")
    formulas = [
        cell
        for row in verification.iter_rows(
            min_row=2, max_row=source.max_row, min_col=19, max_col=23
        )
        for cell in row
        if cell.data_type == "f"
    ]
    if len(formulas) != (source.max_row - 1) * 5:
        raise RuntimeError("No se generaron todas las fórmulas auxiliares de S:W.")
    if any("#REF!" in str(cell.value) for cell in formulas):
        raise RuntimeError("Se detectaron referencias rotas en S:W.")
    verification_book.close()
    source_book.close()
    print(
        f"PegarData validada: A1:R{source.max_row} reemplazado; "
        f"Table2 ampliada a {expected_table_ref}; VBA conservado."
    )
    return destination


def upload_sharepoint_workbook(token: str, item: dict, workbook_path: Path) -> None:
    drive_id = item["parentReference"]["driveId"]
    item_id = item["id"]
    content = workbook_path.read_bytes()
    upload_url = f"{GRAPH_URL}/drives/{drive_id}/items/{item_id}/content"
    headers = {
        **graph_headers(token),
        "Content-Type": "application/octet-stream",
        "If-Match": item["eTag"],
    }

    for attempt in range(1, LOCK_RETRY_ATTEMPTS + 1):
        response = requests.put(upload_url, headers=headers, data=content, timeout=180)
        if response.status_code in {409, 423} and attempt < LOCK_RETRY_ATTEMPTS:
            print(
                f"El libro está abierto o bloqueado temporalmente en SharePoint "
                f"(HTTP {response.status_code}). Reintento "
                f"{attempt}/{LOCK_RETRY_ATTEMPTS - 1} en "
                f"{LOCK_RETRY_SECONDS} segundos..."
            )
            time.sleep(LOCK_RETRY_SECONDS)
            continue
        if response.status_code == 412:
            raise SharePointVersionConflict(
                "El libro cambió en SharePoint durante el proceso."
            )
        response.raise_for_status()
        break

    verification_path = SHAREPOINT_DIR / "Reunion 1-2-3 Test_verificacion.xlsm"
    last_error = None
    for attempt in range(1, 6):
        verification = requests.get(
            upload_url,
            headers=graph_headers(token),
            timeout=180,
        )
        verification.raise_for_status()
        verification_path.write_bytes(verification.content)
        try:
            validate_uploaded_workbook(workbook_path, verification_path)
            print("Libro actualizado y verificado correctamente en SharePoint.")
            return
        except RuntimeError as exc:
            last_error = exc
            if attempt < 5:
                print(f"SharePoint aún procesa el archivo. Verificación {attempt}/5...")
                time.sleep(5)
    raise RuntimeError("No se pudo validar el archivo guardado en SharePoint.") from last_error


def upload_sharepoint_file(token: str, item: dict, filepath: Path) -> None:
    drive_id = item["parentReference"]["driveId"]
    item_id = item["id"]
    content = filepath.read_bytes()
    upload_url = f"{GRAPH_URL}/drives/{drive_id}/items/{item_id}/content"
    headers = {
        **graph_headers(token),
        "Content-Type": "application/octet-stream",
        "If-Match": item["eTag"],
    }

    for attempt in range(1, LOCK_RETRY_ATTEMPTS + 1):
        response = requests.put(upload_url, headers=headers, data=content, timeout=180)
        if response.status_code in {409, 423} and attempt < LOCK_RETRY_ATTEMPTS:
            print(
                f"El archivo está abierto o bloqueado temporalmente en SharePoint "
                f"(HTTP {response.status_code}). Reintento "
                f"{attempt}/{LOCK_RETRY_ATTEMPTS - 1} en "
                f"{LOCK_RETRY_SECONDS} segundos..."
            )
            time.sleep(LOCK_RETRY_SECONDS)
            continue
        if response.status_code == 412:
            raise SharePointVersionConflict("El archivo cambió en SharePoint durante el proceso.")
        response.raise_for_status()
        print(f"Archivo subido correctamente a SharePoint: {filepath.name}")
        break


def upload_test_copy(token: str, item: dict, workbook_path: Path) -> str:
    drive_id = item["parentReference"]["driveId"]
    parent_id = item["parentReference"]["id"]
    filename = "Reunion 1-2-3 Test BOT PRUEBA.xlsm"
    response = requests.put(
        f"{GRAPH_URL}/drives/{drive_id}/items/{parent_id}:/{filename}:/content",
        headers={**graph_headers(token), "Content-Type": "application/octet-stream"},
        data=workbook_path.read_bytes(),
        timeout=180,
    )
    response.raise_for_status()
    web_url = response.json()["webUrl"]
    print(f"COPIA_PRUEBA_URL={web_url}")
    return web_url


def validate_uploaded_workbook(expected_path: Path, remote_path: Path) -> None:
    if vba_hash(remote_path) != vba_hash(expected_path):
        raise RuntimeError("El VBA remoto no coincide con el archivo validado.")

    expected_book = load_workbook(expected_path, keep_vba=True, keep_links=True, data_only=False)
    remote_book = load_workbook(remote_path, keep_vba=True, keep_links=True, data_only=False)
    try:
        if remote_book.sheetnames != expected_book.sheetnames:
            raise RuntimeError("Las hojas del libro remoto no coinciden.")

        expected = expected_book["PegarData"]
        remote = remote_book["PegarData"]
        if (remote.max_row, remote.max_column) != (expected.max_row, expected.max_column):
            raise RuntimeError("Las dimensiones de PegarData no coinciden.")

        for row in range(1, expected.max_row + 1):
            for column in range(1, expected.max_column + 1):
                expected_cell = expected.cell(row=row, column=column)
                remote_cell = remote.cell(row=row, column=column)
                if (
                    remote_cell.value != expected_cell.value
                    or remote_cell.data_type != expected_cell.data_type
                    or remote_cell.number_format != expected_cell.number_format
                    or remote_cell._style != expected_cell._style
                ):
                    raise RuntimeError(
                        f"PegarData remota no coincide en fila {row}, columna {column}."
                    )
    finally:
        expected_book.close()
        remote_book.close()


def sync_report_to_sharepoint(report_path: Path, upload: bool, test_copy: bool = False) -> Path:
    token = graph_token()
    if upload:
        from excel_range_sync import update_pegar_data_range

        item = resolve_sharepoint_item(token)
        update_pegar_data_range(token, item, report_path)
        return report_path

    item = resolve_sharepoint_item(token)
    original = download_sharepoint_workbook(token, item)
    updated = replace_pegar_data(report_path, original)
    if test_copy:
        upload_test_copy(token, item, updated)
    else:
        print("Modo de prueba: el libro fue validado, pero no se subió a SharePoint.")
    return updated
