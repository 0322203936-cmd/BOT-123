from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def copy_cell(source, destination) -> None:
    """Copia contenido y formato sin alterar el valor de la celda."""
    destination.value = source.value
    destination._style = copy(source._style)
    destination._hyperlink = copy(source.hyperlink)
    destination.comment = copy(source.comment)


def format_downloaded_report(source_path: Path) -> Path:
    """Reordena columnas, ajusta txr_orden y activa filtros."""
    workbook = load_workbook(source_path)
    worksheet = workbook.active

    if worksheet.max_column < 19:
        raise RuntimeError(
            f"El reporte tiene {worksheet.max_column} columnas; se esperaban al menos 19."
        )

    expected_header = "txr_orden"
    actual_header = str(worksheet.cell(row=1, column=14).value or "").strip()
    if actual_header != expected_header:
        raise RuntimeError(
            f"La columna N no es {expected_header}; se encontró {actual_header or 'vacía'}."
        )

    hours_header = str(worksheet.cell(row=1, column=19).value or "").strip()
    if hours_header != "Horas":
        raise RuntimeError(
            f"La columna S no es Horas; se encontró {hours_header or 'vacía'}."
        )

    # La macro original mueve N al final del reporte. En la exportación actual,
    # primero se reordena N:R y después se elimina la columna adicional Horas.
    for row in range(1, worksheet.max_row + 1):
        original_cells = [
            copy(worksheet.cell(row=row, column=column))
            for column in range(14, 19)
        ]
        reordered_cells = original_cells[1:] + original_cells[:1]
        for column, source_cell in zip(range(14, 19), reordered_cells):
            copy_cell(source_cell, worksheet.cell(row=row, column=column))

    changed_veronica = 0
    changed_snapdragon = 0
    for row in range(2, worksheet.max_row + 1):
        flower = str(worksheet.cell(row=row, column=5).value or "").strip().upper()
        txr_cell = worksheet.cell(row=row, column=18)

        if flower.startswith("VERONICA"):
            try:
                txr_value = float(str(txr_cell.value).strip())
            except (TypeError, ValueError):
                txr_value = None
            if txr_value not in {8.0, 10.0}:
                txr_cell.value = 8
                changed_veronica += 1
        elif flower.startswith("SNAPDRAGON"):
            try:
                txr_value = float(str(txr_cell.value).strip())
            except (TypeError, ValueError):
                txr_value = None
            if txr_value not in {5.0, 10.0}:
                txr_cell.value = 5
                changed_snapdragon += 1

    worksheet.delete_cols(19, 1)

    last_column = get_column_letter(worksheet.max_column)
    worksheet.auto_filter.ref = f"A1:{last_column}{worksheet.max_row}"

    destination = source_path.with_name(f"{source_path.stem}_formateado.xlsx")
    workbook.save(destination)
    workbook.close()
    print(
        "Ajustes txr_orden: "
        f"Veronica={changed_veronica}, Snapdragon={changed_snapdragon}"
    )
    print(f"Reporte formateado: {destination}")
    return destination
