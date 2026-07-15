import posixpath
import re
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
XML_NS = "http://www.w3.org/XML/1998/namespace"
NS = {"m": MAIN_NS, "r": REL_NS, "p": PKG_REL_NS, "ct": CONTENT_NS}


def qn(namespace: str, name: str) -> str:
    return f"{{{namespace}}}{name}"


def xml_bytes(root) -> bytes:
    return etree.tostring(
        root,
        xml_declaration=True,
        encoding="UTF-8",
        standalone=True,
    )


def relationship_target(base_path: str, target: str) -> str:
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_path), target))


def resolve_paths(files: dict[str, bytes]) -> tuple[str, str, str]:
    workbook_path = "xl/workbook.xml"
    workbook = etree.fromstring(files[workbook_path])
    sheet = workbook.xpath(".//m:sheet[@name='PegarData']", namespaces=NS)
    if len(sheet) != 1:
        raise RuntimeError("No se pudo identificar una sola hoja PegarData.")
    relationship_id = sheet[0].get(qn(REL_NS, "id"))

    workbook_rels_path = "xl/_rels/workbook.xml.rels"
    workbook_rels = etree.fromstring(files[workbook_rels_path])
    relationship = workbook_rels.xpath(
        ".//p:Relationship[@Id=$relationship_id]",
        namespaces=NS,
        relationship_id=relationship_id,
    )
    if len(relationship) != 1:
        raise RuntimeError("No se encontró la relación XML de PegarData.")
    sheet_path = relationship_target(workbook_path, relationship[0].get("Target"))

    sheet_name = posixpath.basename(sheet_path)
    sheet_rels_path = posixpath.join(
        posixpath.dirname(sheet_path), "_rels", f"{sheet_name}.rels"
    )
    sheet_rels = etree.fromstring(files[sheet_rels_path])
    table_relationship = sheet_rels.xpath(
        ".//p:Relationship[contains(@Type, '/table')]", namespaces=NS
    )
    if len(table_relationship) != 1:
        raise RuntimeError("No se encontró una sola tabla relacionada con PegarData.")
    table_path = relationship_target(sheet_path, table_relationship[0].get("Target"))
    return sheet_path, table_path, workbook_rels_path


def column_number(cell_reference: str) -> int:
    match = re.match(r"([A-Z]+)", cell_reference)
    if not match:
        raise RuntimeError(f"Referencia de celda inválida: {cell_reference}")
    number = 0
    for character in match.group(1):
        number = number * 26 + ord(character) - 64
    return number


def source_values(report_path: Path) -> list[list[object]]:
    workbook = load_workbook(report_path, read_only=True, data_only=False)
    worksheet = workbook.active
    if worksheet.max_column != 18:
        workbook.close()
        raise RuntimeError(
            f"El reporte formateado debe tener 18 columnas; tiene {worksheet.max_column}."
        )
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    workbook.close()
    return rows


def add_value_cell(row_element, row: int, column: int, value, style: str | None) -> None:
    if value is None:
        return
    reference = f"{get_column_letter(column)}{row}"
    attributes = {"r": reference}
    if style is not None:
        attributes["s"] = style
    cell = etree.SubElement(row_element, qn(MAIN_NS, "c"), attributes)

    if isinstance(value, bool):
        cell.set("t", "b")
        etree.SubElement(cell, qn(MAIN_NS, "v")).text = "1" if value else "0"
    elif isinstance(value, (int, float)):
        etree.SubElement(cell, qn(MAIN_NS, "v")).text = str(value)
    elif isinstance(value, (date, datetime)):
        etree.SubElement(cell, qn(MAIN_NS, "v")).text = str(to_excel(value))
    else:
        text = str(value)
        if column in {8, 15} and re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
            parsed = datetime.strptime(text, "%Y-%m-%d").date()
            etree.SubElement(cell, qn(MAIN_NS, "v")).text = str(to_excel(parsed))
            return
        cell.set("t", "inlineStr")
        inline = etree.SubElement(cell, qn(MAIN_NS, "is"))
        text_node = etree.SubElement(inline, qn(MAIN_NS, "t"))
        if text != text.strip():
            text_node.set(qn(XML_NS, "space"), "preserve")
        text_node.text = text


def add_formula_cell(
    row_element,
    row: int,
    column: int,
    formula: str,
    cached_value,
    style: str | None,
) -> None:
    attributes = {"r": f"{get_column_letter(column)}{row}"}
    if style is not None:
        attributes["s"] = style
    if isinstance(cached_value, str):
        attributes["t"] = "str"
    cell = etree.SubElement(row_element, qn(MAIN_NS, "c"), attributes)
    etree.SubElement(cell, qn(MAIN_NS, "f")).text = formula
    value = etree.SubElement(cell, qn(MAIN_NS, "v"))
    value.text = "" if cached_value is None else str(cached_value)


def formula_cells(row_element, row: int, values: list[object], styles: dict[int, str | None]) -> None:
    flower = "" if values[4] is None else str(values[4])
    color = "" if values[5] is None else str(values[5])
    txr = "" if values[17] is None else str(values[17])
    formulas = {
        19: (
            '+Table2[[#This Row],[FLOR]]&" "&"X"&" " & Table2[[#This Row],[txr_orden]]',
            f"{flower} X {txr}",
        ),
        20: ("Table2[[#This Row],[txr_orden]]", values[17]),
        21: (
            '+Table2[[#This Row],[FLOR]] & " " &Table2[[#This Row],[Flor Color]]',
            f"{flower} {color}",
        ),
        22: (
            '+Table2[[#This Row],[FLOR]] & " " & Table2[[#This Row],[Flor Color]]&"X"&" " & Table2[[#This Row],[TxR2]]',
            f"{flower} {color}X {txr}",
        ),
        23: (f'E{row} & " " & " X " & T{row}', f"{flower}  X {txr}"),
    }
    for column, (formula, cached_value) in formulas.items():
        add_formula_cell(row_element, row, column, formula, cached_value, styles[column])


def patch_sheet(sheet_root, rows: list[list[object]]) -> None:
    sheet_data = sheet_root.find(qn(MAIN_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("PegarData no contiene sheetData.")
    row_elements = {int(row.get("r")): row for row in sheet_data.findall(qn(MAIN_NS, "row"))}
    header = row_elements.get(1)
    template = row_elements.get(2)
    if header is None or template is None:
        raise RuntimeError("PegarData no contiene filas de encabezado y plantilla.")

    header_values = [cell for cell in rows[0]]
    expected_headers = [
        "Distribuidor", "Cust Name", "No.", "CATEGORIA", "FLOR", "Flor Color",
        "Descripcion", "Load Date", "#cajas", "Prod", "Term", "#Ramos",
        "Ramos/Cajas", "Total Tallos", "Fecha Produccion", "Caja", "Linea",
        "txr_orden",
    ]
    if header_values != expected_headers:
        raise RuntimeError("Los encabezados del reporte formateado no coinciden con PegarData.")

    styles: dict[int, str | None] = {}
    for cell in template.findall(qn(MAIN_NS, "c")):
        styles[column_number(cell.get("r"))] = cell.get("s")
    if any(column not in styles for column in range(1, 24)):
        raise RuntimeError("No se pudieron obtener todos los estilos de la fila plantilla.")

    new_last_row = len(rows)
    for row_number in range(2, new_last_row + 1):
        row_element = row_elements.get(row_number)
        if row_element is None:
            row_element = etree.Element(qn(MAIN_NS, "row"), {"r": str(row_number)})
            sheet_data.append(row_element)
            row_elements[row_number] = row_element
        preserved = [
            deepcopy(cell)
            for cell in row_element.findall(qn(MAIN_NS, "c"))
            if column_number(cell.get("r")) > 23
        ]
        for cell in list(row_element.findall(qn(MAIN_NS, "c"))):
            row_element.remove(cell)
        row_element.set("r", str(row_number))
        row_element.set("spans", "1:24" if preserved else "1:23")

        values = rows[row_number - 1]
        for column, value in enumerate(values, start=1):
            add_value_cell(row_element, row_number, column, value, styles[column])
        formula_cells(row_element, row_number, values, styles)
        for cell in sorted(preserved, key=lambda item: column_number(item.get("r"))):
            row_element.append(cell)

    for row_number, row_element in list(row_elements.items()):
        if row_number <= new_last_row:
            continue
        preserved = [
            deepcopy(cell)
            for cell in row_element.findall(qn(MAIN_NS, "c"))
            if column_number(cell.get("r")) > 23
        ]
        if not preserved:
            sheet_data.remove(row_element)
            continue
        for cell in list(row_element.findall(qn(MAIN_NS, "c"))):
            row_element.remove(cell)
        for cell in preserved:
            row_element.append(cell)

    ordered_rows = sorted(sheet_data.findall(qn(MAIN_NS, "row")), key=lambda item: int(item.get("r")))
    for row_element in list(sheet_data):
        sheet_data.remove(row_element)
    for row_element in ordered_rows:
        sheet_data.append(row_element)

    dimension = sheet_root.find(qn(MAIN_NS, "dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:X{max(new_last_row, 8)}")


def patch_table(table_root, last_row: int) -> None:
    if table_root.get("name") != "Table2" or table_root.get("displayName") != "Table2":
        raise RuntimeError("La tabla de PegarData no es Table2.")
    reference = f"A1:W{last_row}"
    table_root.set("ref", reference)
    auto_filter = table_root.find(qn(MAIN_NS, "autoFilter"))
    if auto_filter is None:
        raise RuntimeError("Table2 no contiene autoFilter.")
    auto_filter.set("ref", reference)


def patch_workbook(workbook_root) -> None:
    calc_properties = workbook_root.find(qn(MAIN_NS, "calcPr"))
    if calc_properties is None:
        calc_properties = etree.SubElement(workbook_root, qn(MAIN_NS, "calcPr"))
    calc_properties.set("calcMode", "auto")
    calc_properties.set("fullCalcOnLoad", "1")
    calc_properties.set("forceFullCalc", "1")


def remove_calc_chain(files: dict[str, bytes], workbook_rels_path: str) -> None:
    files.pop("xl/calcChain.xml", None)
    relationships = etree.fromstring(files[workbook_rels_path])
    for relationship in relationships.xpath(
        ".//p:Relationship[contains(@Type, '/calcChain')]", namespaces=NS
    ):
        relationship.getparent().remove(relationship)
    files[workbook_rels_path] = xml_bytes(relationships)

    content_types = etree.fromstring(files["[Content_Types].xml"])
    for override in content_types.xpath(
        ".//ct:Override[@PartName='/xl/calcChain.xml']", namespaces=NS
    ):
        override.getparent().remove(override)
    files["[Content_Types].xml"] = xml_bytes(content_types)


def patch_xlsm(report_path: Path, source_xlsm: Path, destination: Path) -> Path:
    rows = source_values(report_path)
    with ZipFile(source_xlsm) as source_archive:
        infos = source_archive.infolist()
        files = {info.filename: source_archive.read(info.filename) for info in infos}

    sheet_path, table_path, workbook_rels_path = resolve_paths(files)
    sheet_root = etree.fromstring(files[sheet_path])
    table_root = etree.fromstring(files[table_path])
    workbook_root = etree.fromstring(files["xl/workbook.xml"])

    patch_sheet(sheet_root, rows)
    patch_table(table_root, len(rows))
    patch_workbook(workbook_root)
    files[sheet_path] = xml_bytes(sheet_root)
    files[table_path] = xml_bytes(table_root)
    files["xl/workbook.xml"] = xml_bytes(workbook_root)
    remove_calc_chain(files, workbook_rels_path)

    destination.parent.mkdir(parents=True, exist_ok=True)
    with ZipFile(destination, "w", compression=ZIP_DEFLATED, allowZip64=True) as output:
        written = set()
        for info in infos:
            if info.filename not in files:
                continue
            output.writestr(info, files[info.filename])
            written.add(info.filename)
        for name, content in files.items():
            if name not in written:
                output.writestr(name, content)
    validate_unchanged_parts(
        source_xlsm,
        destination,
        changed_parts={
            sheet_path,
            table_path,
            "xl/workbook.xml",
            workbook_rels_path,
            "[Content_Types].xml",
            "xl/calcChain.xml",
        },
    )
    return destination


def validate_unchanged_parts(
    source_xlsm: Path,
    destination: Path,
    changed_parts: set[str],
) -> None:
    with ZipFile(source_xlsm) as source, ZipFile(destination) as output:
        if source.testzip() is not None or output.testzip() is not None:
            raise RuntimeError("El paquete XLSM contiene una entrada ZIP dañada.")
        source_names = set(source.namelist())
        output_names = set(output.namelist())
        expected_names = source_names - {"xl/calcChain.xml"}
        if output_names != expected_names:
            raise RuntimeError("La lista de componentes del XLSM cambió inesperadamente.")
        for name in expected_names - changed_parts:
            if source.read(name) != output.read(name):
                raise RuntimeError(f"El componente protegido {name} fue modificado.")
