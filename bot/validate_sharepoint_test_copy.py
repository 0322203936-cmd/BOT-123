from pathlib import Path
from urllib.parse import quote
from zipfile import ZipFile

import requests
from openpyxl import load_workbook

from sharepoint_sync import GRAPH_URL, graph_headers, graph_token, resolve_sharepoint_item


def run() -> None:
    token = graph_token()
    main_item = resolve_sharepoint_item(token)
    drive_id = main_item["parentReference"]["driveId"]
    parent_id = main_item["parentReference"]["id"]
    filename = "Reunion 1-2-3 Test BOT PRUEBA.xlsm"
    encoded_name = quote(filename, safe="")
    item_url = f"{GRAPH_URL}/drives/{drive_id}/items/{parent_id}:/{encoded_name}"

    metadata = requests.get(item_url, headers=graph_headers(token), timeout=30)
    metadata.raise_for_status()
    item = metadata.json()

    output_dir = Path("artifacts/test-copy-validation")
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = requests.get(
        f"{GRAPH_URL}/drives/{drive_id}/items/{item['id']}/content",
        headers=graph_headers(token),
        timeout=180,
    )
    workbook.raise_for_status()
    workbook_path = output_dir / filename
    workbook_path.write_bytes(workbook.content)

    with ZipFile(workbook_path) as archive:
        names = set(archive.namelist())
        if "xl/vbaProject.bin" not in names:
            raise RuntimeError("La copia de prueba perdio el proyecto VBA.")

    book = load_workbook(workbook_path, read_only=False, data_only=False, keep_vba=True)
    try:
        if "PegarData" not in book.sheetnames:
            raise RuntimeError("No existe la hoja PegarData.")
        sheet = book["PegarData"]
        if "Table2" not in sheet.tables:
            raise RuntimeError("No existe la tabla Table2 en PegarData.")

        table_ref = sheet.tables["Table2"].ref
        expected_ref = f"A1:W{sheet.max_row}"
        if table_ref != expected_ref:
            raise RuntimeError(
                f"Rango inesperado de Table2: {table_ref}; esperado: {expected_ref}."
            )

        formulas = []
        ref_errors = []
        for row in sheet.iter_rows(min_row=2, min_col=19, max_col=23):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(cell.coordinate)
                    if "#REF!" in cell.value.upper():
                        ref_errors.append(cell.coordinate)

        expected_formulas = max(sheet.max_row - 1, 0) * 5
        if len(formulas) != expected_formulas:
            raise RuntimeError(
                f"Cantidad inesperada de formulas S:W: {len(formulas)}; "
                f"esperadas: {expected_formulas}."
            )
        if ref_errors:
            raise RuntimeError(
                f"Se encontraron formulas con #REF!: {', '.join(ref_errors[:10])}"
            )
    finally:
        book.close()

    print(
        f"ESTRUCTURA_XLSM_OK bytes={workbook_path.stat().st_size} "
        f"tabla={table_ref} formulas={len(formulas)} ref_errors={len(ref_errors)} "
        f"vba=conservado item={item['id']}"
    )

    pdf = requests.get(
        f"{GRAPH_URL}/drives/{drive_id}/items/{item['id']}/content?format=pdf",
        headers=graph_headers(token),
        timeout=300,
    )
    if pdf.ok:
        pdf_path = output_dir / "Reunion BOT PRUEBA.pdf"
        pdf_path.write_bytes(pdf.content)
        print(f"OFFICE_CONVERSION_OK pdf={pdf_path.stat().st_size}")
    elif pdf.status_code == 406:
        print("PDF_CONVERSION_NO_DISPONIBLE status=406 (no es un error del XLSM)")
    else:
        pdf.raise_for_status()


if __name__ == "__main__":
    run()
