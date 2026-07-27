import sys
import openpyxl

from sharepoint_sync import (
    graph_token,
    resolve_sharepoint_item_by_url,
    GRAPH_URL,
    graph_headers
)
from excel_range_sync import graph_request

REQ_PROY_URL = "https://pacificafarms.sharepoint.com/:x:/r/sites/requerimientovsproyeccion/_layouts/15/Doc.aspx?sourcedoc=%7BB6111299-1373-4717-A2B7-3D507AD77A8A%7D&file=Requerimiento%20vs%20proyeccion.xlsm&action=default&mobileredirect=true"
LOCAL_FILE = r"C:\Users\Yisus\Desktop\Atajos Globales\Requerimiento vs proyeccion Test.xlsm"

def col_letter_to_index(letter):
    idx = 0
    for char in letter.upper():
        idx = idx * 26 + (ord(char) - ord('A') + 1)
    return idx - 1

def read_local():
    print(f"Leyendo archivo local: {LOCAL_FILE}")
    wb = openpyxl.load_workbook(LOCAL_FILE, data_only=True)
    ws = wb["DataProy"]
    data = []
    for row in range(1, ws.max_row + 1):
        desc = ws.cell(row=row, column=8).value
        if str(desc).strip().upper() == "CORTE":
            flor = ws.cell(row=row, column=6).value
            color = ws.cell(row=row, column=7).value
            tallos = ws.cell(row=row, column=15).value
            semana = ws.cell(row=row, column=19).value
            data.append({
                "row": row,
                "flor": flor if flor is not None else "",
                "color": color if color is not None else "",
                "tallos": tallos if tallos is not None else "",
                "semana": semana if semana is not None else ""
            })
    wb.close()
    return data

def read_cloud():
    print("Obteniendo datos de SharePoint...")
    token = graph_token()
    item_req = resolve_sharepoint_item_by_url(token, REQ_PROY_URL)
    drive_id = item_req["parentReference"]["driveId"]
    workbook_url = f"{GRAPH_URL}/drives/{drive_id}/items/{item_req['id']}/workbook"
    headers = {**graph_headers(token), "Content-Type": "application/json"}
    
    session = graph_request("POST", f"{workbook_url}/createSession", headers, json={"persistChanges": False}, timeout=60).json()
    session_headers = {**headers, "workbook-session-id": session["id"]}
    
    try:
        used_range_res = graph_request("GET", f"{workbook_url}/worksheets/DataProy/usedRange", session_headers, timeout=120).json()
        address = used_range_res.get("address", "")
        values = used_range_res.get("values", [])
        
        import re
        match = re.search(r'!([A-Za-z]+)(\d+)', address)
        start_col_str = match.group(1).upper() if match else "A"
        start_row = int(match.group(2)) if match else 1
        
        start_col_idx = col_letter_to_index(start_col_str)
        col_f = col_letter_to_index("F") - start_col_idx
        col_g = col_letter_to_index("G") - start_col_idx
        col_h = col_letter_to_index("H") - start_col_idx
        col_o = col_letter_to_index("O") - start_col_idx
        col_s = col_letter_to_index("S") - start_col_idx
        
        data = []
        for idx, row_data in enumerate(values):
            if len(row_data) > col_h and col_h >= 0:
                desc = str(row_data[col_h]).strip().upper()
                if desc == "CORTE":
                    flor = row_data[col_f] if len(row_data) > col_f else ""
                    color = row_data[col_g] if len(row_data) > col_g else ""
                    tallos = row_data[col_o] if len(row_data) > col_o else ""
                    semana = row_data[col_s] if len(row_data) > col_s else ""
                    
                    data.append({
                        "row": start_row + idx,
                        "flor": flor if flor is not None else "",
                        "color": color if color is not None else "",
                        "tallos": tallos if tallos is not None else "",
                        "semana": semana if semana is not None else ""
                    })
        return data
    finally:
        try:
            graph_request("POST", f"{workbook_url}/closeSession", session_headers, timeout=30)
        except Exception:
            pass

def main():
    local_data = read_local()
    cloud_data = read_cloud()
    
    print(f"\n--- COMPARACIÓN DE RESULTADOS ---")
    print(f"Filas de CORTE encontradas: LOCAL = {len(local_data)} | NUBE = {len(cloud_data)}")
    
    limit = min(len(local_data), len(cloud_data))
    diferencias = 0
    for i in range(limit):
        loc = local_data[i]
        cld = cloud_data[i]
        
        diffs = []
        if str(loc['flor']).strip() != str(cld['flor']).strip(): diffs.append(f"Flor (L:{loc['flor']} vs N:{cld['flor']})")
        if str(loc['color']).strip() != str(cld['color']).strip(): diffs.append(f"Color (L:{loc['color']} vs N:{cld['color']})")
        # For tallos, handle float vs int formatting by comparing float values if possible
        try:
            if float(loc['tallos'] or 0) != float(cld['tallos'] or 0): diffs.append(f"Tallos (L:{loc['tallos']} vs N:{cld['tallos']})")
        except:
            if str(loc['tallos']).strip() != str(cld['tallos']).strip(): diffs.append(f"Tallos (L:{loc['tallos']} vs N:{cld['tallos']})")
            
        try:
            if float(loc['semana'] or 0) != float(cld['semana'] or 0): diffs.append(f"Semana (L:{loc['semana']} vs N:{cld['semana']})")
        except:
            if str(loc['semana']).strip() != str(cld['semana']).strip(): diffs.append(f"Semana (L:{loc['semana']} vs N:{cld['semana']})")
        
        if diffs:
            diferencias += 1
            print(f"Diferencia fila Nube {cld['row']} / Local {loc['row']}: {', '.join(diffs)}")
            
    if diferencias == 0 and len(local_data) == len(cloud_data):
        print("\n✅ ¡EXITO! El archivo automatizado en la nube (SharePoint) está EXACTAMENTE IGUAL que tu archivo de base (Local).")
    elif diferencias == 0:
        print(f"\n⚠️ Casi perfecto. No hay diferencias en el texto, pero hay distinta cantidad de filas totales (Local: {len(local_data)}, Nube: {len(cloud_data)}).")
    else:
        print(f"\n❌ Se encontraron {diferencias} filas con diferencias.")

if __name__ == '__main__':
    main()
